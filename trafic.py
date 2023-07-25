import pandas as pd 
import numpy as np
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill

# input start
fd = input('Input file directory:')
ym = input('Input year & month (Ex: 202301--> year=2023 & month=01):')
date = input('input date(two digits):')
tlc = int(input('Input Site Number:'))
site_name = input('Input Site Name (Ex:Lonsdale Street/Foster St):')
out_path = input('Output file directory:')
out_file = input('Output file name:')
# input end

# input approaches
data_north = {}
data_south = {}
data_east = {}
data_west = {}
approach = list(map(str,input('Input approaches(eg: North,South...): ').split(',')))
if 'North' in approach:
    opt = 'y'
    while opt=='y':
        key = input('North approach specific turns: ')
        value = list(map(int,input('Turn specific detector no: ').split(',')))
        data_north[key] = value
        opt = input('Any specific turns and detector numbers left(y/n): ')
if 'South' in approach:
    opt = 'y'
    while opt=='y':
        key = input('South approach specific turns: ')
        value = list(map(int,input('Turn specific detector no: ').split(',')))
        data_south[key] = value
        opt = input('Any specific turns and detector numbers left(y/n): ')
if 'East' in approach:
    opt = 'y'
    while opt=='y':
        key = input('East approach specific turns: ')
        value = list(map(int,input('Turn specific detector no:: ').split(',')))
        data_east[key] = value
        opt = input('Any specific turns and detector numbers left(y/n): ')
if 'West' in approach:
    opt = 'y'
    while opt=='y':
        key = input('West approach specific turns:')
        value = list(map(int,input('Turn specific detector no: ').split(',')))
        data_west[key] = value
        opt = input('Any specific turns and detector numbers left(y/n): ')
# input approaches-->end




# file input to pandas
file_name = 'VSDATA_'+ym
csv_name = 'VSDATA_'+ym+date+'.csv'
file = pd.read_csv(fd+'/'+file_name+'/'+csv_name)
# file input to pandas -->end

#finding data according to tlc number
tlc_file = file[(file.NB_SCATS_SITE == tlc)]

# Define a function to format the time intervals in the desired format
def format_time_interval(interval):
    hour = interval // 4
    minute_start = (interval % 4) * 15
    minute_end = ((interval % 4) * 15 + 15) % 60
    hour_next = (hour + 1) % 24 if minute_end == 0 else hour
    
    if hour == 24:
        hour = 0
    if hour_next == 24:
        hour_next = 0
    
    return f'{hour:02d}{minute_start:02d}-{hour_next:02d}{minute_end:02d}'

# Get the column names excluding 'NB_DETECTOR'
columns = [col for col in tlc_file.columns.tolist() if col != 'NB_DETECTOR']

# Rename the columns using the formatted time intervals
new_columns = []
for col in columns:
    try:
        interval = int(col[1:])
        new_columns.append(format_time_interval(interval))
    except ValueError:
        new_columns.append(col)
        
tlc_file = tlc_file.rename(columns=dict(zip(columns, new_columns)))


# Combine all values from the input data dictionaries
all_values = list(data_north.values()) + list(data_south.values()) + list(data_east.values()) + list(data_west.values())
num_rows = sum(len(lst) for lst in all_values)

# Get the lengths of all lists in all_values
lengths = [len(lst) for lst in all_values]

# Find the largest length
largest_length = max(lengths)

# Create the list of detector fields
detector_fields = [f"Detectors{i+1}" for i in range(largest_length)]

# Create a list to store the rows
rows = []

# Iterate over the approach values
for appr in approach:
    if appr == 'North':
        movement_data = data_north
    elif appr == 'South':
        movement_data = data_south
    elif appr == 'East':
        movement_data = data_east
    elif appr == 'West':
        movement_data = data_west
    else:
        continue
    
    # Iterate over the movement and detector values
    for movement, detectors in movement_data.items():
        # Create a dictionary for the row data
        row_data = {
            'Site': tlc,
            'Approach': appr,
            'Movement': movement
        }
        
        # Add the detector values to the row data
        for i, detector in enumerate(detectors):
            row_data[detector_fields[i]] = detector
        
        # Append the row dictionary to the list
        rows.append(row_data)

# Create the DataFrame from the list of rows
df1 = pd.DataFrame(rows)

# input time and interval
start_time = input('Input the Start time(ex:HHMM-->0330 or 1515):')
finish_time = input('Input the Finish time(ex:HHMM-->0330 or 1515):')
interval = int(input('Input the time interval(in minutes):'))
# input time and interval -->end

num_rows = sum(len(lst) for lst in all_values)
# Convert start_time and finish_time to integers
start_hour = int(start_time[:2])
start_minute = int(start_time[2:])
finish_hour = int(finish_time[:2])
finish_minute = int(finish_time[2:])

# Calculate the number of intervals
num_intervals = int(((finish_hour * 60 + finish_minute) - (start_hour * 60 + start_minute)) / interval)

# Create the time ranges
time_ranges = []
for i in range(num_intervals):
    start = f'{start_hour:02d}{start_minute:02d}'
    start_minute += interval
    if start_minute >= 60:
        start_hour += 1
        start_minute -= 60
    end = f'{start_hour:02d}{start_minute:02d}'
    time_ranges.append(f'{start} - {end}')

# Create the DataFrame with empty values
data = np.empty((num_rows, num_intervals), dtype=object)
data.fill(np.nan)
df2 = pd.DataFrame(data, columns=time_ranges)

# Concatenate the two DataFrames horizontally
df1 = df1.join(df2)
# Read the second CSV file
df2 = tlc_file.copy()

# Iterate over each row in df1
for index, row in df1.iterrows():
    time_columns = df1.columns[5:]  # Get the time range columns
    
    # Iterate over each time range column in df1
    for column in time_columns:
        if 'Detectors' in column:
            continue  # Skip the detector columns
        
        time_range = column.split('-')
        start_time = int(time_range[0])
        end_time = int(time_range[1])
        if end_time == 0:
            end_time = 2400
        
        sum_value = 0  # Initialize the sum of detectors for the time range
        
        # Iterate over each detector column in df1
        for detector_column in detector_fields:
            detector_number = row[detector_column]  # Get the detector number
            
            # Filter the columns in df2 based on time range and detector number
            filtered_columns = [col for col in df2.columns[2:] if col not in ['NB_DETECTOR', 'NM_REGION', 'CT_RECORDS', 'QT_VOLUME_24HOUR', 'CT_ALARM_24HOUR', 'QT_INTERVAL_COUNT', 'NB_SCATS_SITE'] and start_time <= int(col.split('-')[0]) < end_time]
            
            # Filter the values in df2 based on detector number
            filtered_values = df2.loc[df2['NB_DETECTOR'] == detector_number, filtered_columns]
            
            # Calculate the sum of the filtered values
            value_sum = filtered_values.values.sum()
            
            sum_value += value_sum  # Add the sum to the sum_value
        
        # Update the corresponding cell in df1 with the calculated sum
        df1.at[index, column] = sum_value

# Create a new workbook and select the active sheet
workbook = Workbook()
sheet = workbook.active

# Transfer data from DataFrame to worksheet
for r in dataframe_to_rows(df1, index=False, header=True):
    sheet.append(r)

# Define the number of empty rows to insert
empty_rows = 4

# Get the index of the starting row of the existing data set
start_row = 1  # Replace with the appropriate starting row index

# Shift existing rows down
sheet.insert_rows(start_row, amount=empty_rows)

# Define the row index for the heading row
heading_row = 5

# Define the fill color for the heading row
fill_color = "ADD8E6"  # Replace with the desired color code

# Iterate over the cells in the heading row and apply the fill color
for cell in sheet[heading_row]:
    cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

year = int(ym)//100
month = int(ym)%100
sheet['A1'].value = 'Traffic Volume'
sheet['A2'].value = 'Site:TCS' + f'{tlc:04d}' + '(' + site_name + ')'
sheet['A3'].value = 'Date:' + date + '/' + f'{month:02d}' + '/' + str(year)

# Save the workbook
workbook.save(out_path+'/'+out_file+'.xlsx')

# Close the workbook
workbook.close()

print("Data transfer completed successfully!")
